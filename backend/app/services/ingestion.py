"""Extract uploaded documents to normalized Markdown + content hash."""
from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
from pathlib import Path

from app.config import settings

logger = logging.getLogger("ba-agent.ingestion")

SUPPORTED = {".pdf", ".docx", ".txt", ".md", ".xlsx", ".csv"}


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()


def safe_slug(name: str) -> str:
    name = re.sub(r"[<>:\"/\\|?*]+", "-", name).strip()
    name = re.sub(r"\s+", "-", name)
    return name[:120] or "file"


def workspace_context_dir(project_id: str) -> Path:
    d = settings.workspaces_dir / project_id / "context"
    d.mkdir(parents=True, exist_ok=True)
    return d


def extract_to_markdown(path: Path) -> str:
    """Return normalized Markdown text for a supported file. Never raises on content issues."""
    ext = path.suffix.lower()
    try:
        if ext in (".txt", ".md"):
            return path.read_text(encoding="utf-8", errors="ignore")
        if ext == ".pdf":
            return _extract_pdf(path)
        if ext == ".docx":
            return _extract_docx(path)
        if ext == ".xlsx":
            return _extract_xlsx(path)
        if ext == ".csv":
            return _extract_csv(path)
    except Exception as e:
        logger.error("extract failed for %s: %s", path.name, e)
        return f"*(Could not extract content from {path.name}: {e})*"
    return f"*(Unsupported file type: {path.name})*"


def _extract_pdf(path: Path) -> str:
    try:
        import pdfplumber

        parts: list[str] = []
        with pdfplumber.open(str(path)) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                txt = page.extract_text() or ""
                parts.append(f"### Page {i}\n\n{txt}")
                for table in page.extract_tables() or []:
                    parts.append(_table_to_md(table))
        return "\n\n".join(parts).strip() or "*(No extractable text)*"
    except Exception:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        return "\n\n".join((p.extract_text() or "") for p in reader.pages).strip()


def _extract_docx(path: Path) -> str:
    from docx import Document as Docx

    doc = Docx(str(path))
    lines: list[str] = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style = (para.style.name or "").lower()
        if style.startswith("heading"):
            lines.append(f"## {text}")
        else:
            lines.append(text)
    for table in doc.tables:
        rows = [[c.text for c in row.cells] for row in table.rows]
        if rows:
            lines.append(_table_to_md(rows))
    return "\n\n".join(lines).strip()


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        rows = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(cell.strip() for cell in r)]
        if not rows:
            continue
        parts.append(f"### Sheet: {ws.title}\n\n{_table_to_md(rows)}")
    return "\n\n".join(parts).strip()


def _extract_csv(path: Path) -> str:
    text = path.read_text(encoding="utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    return _table_to_md(rows)


def _table_to_md(rows: list[list]) -> str:
    rows = [[str(c) if c is not None else "" for c in r] for r in rows if r]
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header = rows[0]
    body = rows[1:]
    out = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
    for r in body:
        out.append("| " + " | ".join(cell.replace("\n", " ") for cell in r) + " |")
    return "\n".join(out)


def rebuild_index(project_id: str, docs: list[tuple[str, str]]) -> None:
    """Write context/INDEX.md listing all active docs. docs = [(filename, extracted_relpath)]."""
    ctx = workspace_context_dir(project_id)
    lines = ["# Document Index", "", "The following client documents are available in this folder:", ""]
    for filename, rel in docs:
        lines.append(f"- **{filename}** → `{rel}`")
    (ctx / "INDEX.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
